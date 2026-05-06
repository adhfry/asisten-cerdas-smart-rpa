import openpyxl
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException
from selenium.webdriver.chrome.options import Options
import time
import random
import os
import glob
import sys
import json
from datetime import datetime
import urllib.request
import urllib.error

AUTO_PRINT_DELAY_SPP_MIN = 5
AUTO_PRINT_DELAY_SPP_MAX = 10
AUTO_PRINT_DELAY_FKPP_MIN = 10
AUTO_PRINT_DELAY_FKPP_MAX = 15
AUTO_PRINT_POST_DELAY_SPP = 2
AUTO_PRINT_POST_DELAY_FKPP = 3
ENABLE_KIOSK_PRINTING = True
PRINTER_NAME = "EPSON L120 Series"
NOTIFY_URL = "https://api.silakes.labkesdasumenep.id/api/bot/v1/send-group-message"
NOTIFY_SECRET_KEY = "kesehatanNo1@"

# ====================================================================
# FUNGSI UTILITAS & UI TERMINAL (INTERAKTIF)
# ====================================================================

def get_key():
    if os.name == 'nt':
        import msvcrt
        key = msvcrt.getch()
        if key == b'\xe0':
            key = msvcrt.getch()
            if key == b'H': return 'UP'
            elif key == b'P': return 'DOWN'
        elif key == b'\r': return 'ENTER'
        elif key == b'\x03': raise KeyboardInterrupt
        return key.decode('utf-8', 'ignore')
    else:
        import tty, termios
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setraw(sys.stdin.fileno())
            ch = sys.stdin.read(1)
            if ch == '\x1b':
                ch = sys.stdin.read(2)
                if ch == '[A': return 'UP'
                elif ch == '[B': return 'DOWN'
            elif ch == '\r' or ch == '\n': return 'ENTER'
            elif ch == '\x03': raise KeyboardInterrupt
            return ch
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)

def menu_interaktif(pilihan, judul="Pilih salah satu:"):
    indeks_terpilih = 0
    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        print("="*60)
        print(f" {judul} ")
        print(" (Gunakan panah Atas/Bawah, lalu tekan ENTER)")
        print("="*60)
        
        for i, opsi in enumerate(pilihan):
            if i == indeks_terpilih:
                print(f" [*] {opsi}")
            else:
                print(f" [ ] {opsi}")
                
        key = get_key()
        if key == 'UP':
            indeks_terpilih = (indeks_terpilih - 1) % len(pilihan)
        elif key == 'DOWN':
            indeks_terpilih = (indeks_terpilih + 1) % len(pilihan)
        elif key == 'ENTER':
            os.system('cls' if os.name == 'nt' else 'clear')
            return indeks_terpilih, pilihan[indeks_terpilih]

def setup_folder_excel():
    nama_folder = "FILE EXCEL DISINI"
    if not os.path.exists(nama_folder):
        os.makedirs(nama_folder)
        print(f"\n[INFO] Folder '{nama_folder}' telah dibuat.")
        print("Silakan pindahkan file Excel (data pasien) Anda ke dalam folder tersebut.")
        input("Tekan ENTER jika file sudah dipindahkan...")
    
    while True:
        list_file = glob.glob(os.path.join(nama_folder, "*.xlsx"))
        if not list_file:
            print(f"\n[!] Tidak ada file .xlsx ditemukan di dalam folder '{nama_folder}'.")
            input("Silakan masukkan file, lalu tekan ENTER untuk mengecek kembali...")
            continue
        
        nama_file_saja = [os.path.basename(f) for f in list_file]
        idx_file, file_terpilih = menu_interaktif(nama_file_saja, "Pilih File Excel yang akan diproses:")
        path_file = os.path.join(nama_folder, file_terpilih)
        
        try:
            wb = openpyxl.load_workbook(path_file)
            list_sheet = wb.sheetnames
            idx_sheet, sheet_terpilih = menu_interaktif(list_sheet, f"Pilih Sheet dari file {file_terpilih}:")
            
            os.system('cls' if os.name == 'nt' else 'clear')
            print("="*60)
            print(" KONFIRMASI DATA")
            print("="*60)
            print(f"File  : {file_terpilih}")
            print(f"Sheet : {sheet_terpilih}")
            print("="*60)
            konfirmasi = input("Apakah data ini sudah benar? (Y/Enter = Ya, N = Pilih Ulang): ").strip().lower()
            if konfirmasi == 'n':
                continue
            
            return path_file, sheet_terpilih, wb
        except Exception as e:
            print(f"Error membaca file: {e}")
            input("Tekan ENTER untuk mencoba lagi...")

# ====================================================================
# FUNGSI SELENIUM
# ====================================================================

def ketik_seperti_manusia(elemen, teks):
    for huruf in teks:
        elemen.send_keys(huruf)
        time.sleep(random.uniform(0.01, 0.05))

def tunggu_loading_pace(driver):
    try:
        WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".pace-active")))
        time.sleep(0.5) 
    except:
        pass

def cek_sesi_berakhir(driver):
    try:
        modal_sesi = driver.find_elements(By.XPATH, "//div[contains(text(), 'Sesi Anda sudah berakhir')]")
        if modal_sesi and modal_sesi[0].is_displayed():
            print("\n" + "!"*60)
            print("⚠️ PERINGATAN: SESI PCARE ANDA TELAH BERAKHIR!")
            print("Bot akan menjeda proses. Silakan lakukan hal berikut:")
            print("1. Klik OK pada pop-up di browser.")
            print("2. Login kembali dan selesaikan CAPTCHA.")
            print("3. Pastikan Anda sudah kembali ke Dashboard utama PCare.")
            print("!"*60)
            input("\n[TEKAN ENTER DI SINI JIKA ANDA SUDAH LOGIN KEMBALI KE DASHBOARD]...")
            return True
    except:
        pass
    return False

def pilih_select2(driver, elemen_id_asli, nilai_pencarian):
    try:
        container = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, f"//select[@id='{elemen_id_asli}']/following-sibling::span[contains(@class, 'select2-container')]"))
        )
        container.click()
        time.sleep(0.5)
        
        opsi = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, f"//ul[contains(@id, 'select2-{elemen_id_asli}-results')]//li[contains(text(), '{nilai_pencarian}')]"))
        )
        opsi.click()
        time.sleep(0.2)
    except Exception as e:
        print(f"   [!] Gagal memilih dropdown {elemen_id_asli}: {str(e)[:30]}")

def elemen_terlihat(driver, by, selector):
    try:
        elemen = driver.find_element(by, selector)
        return elemen.is_displayed()
    except Exception:
        return False

def safe_click(driver, by, selector, timeout=8, scroll=True):
    try:
        elemen = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, selector)))
        if scroll:
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemen)
            time.sleep(0.2)
        elemen.click()
        return True
    except Exception:
        try:
            elemen = driver.find_element(by, selector)
            if scroll:
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elemen)
                time.sleep(0.2)
            driver.execute_script("arguments[0].click();", elemen)
            return True
        except Exception as e:
            print(f"⚠️ Gagal klik elemen {selector}: {str(e)[:30]}")
            return False

def buka_tab(driver, tab_id, timeout=8):
    if not safe_click(driver, By.XPATH, f"//a[@href='#{tab_id}']"):
        return False
    try:
        WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, tab_id)))
        return True
    except Exception:
        return False

def daftar_mengandung(daftar, kata_kunci):
    kata = str(kata_kunci).upper()
    return any(kata in str(item).upper() for item in daftar)

def pilih_opsi_dropdown(driver, elemen_id, kata_kunci):
    try:
        select_el = driver.find_element(By.ID, elemen_id)
        sel = Select(select_el)
        for opt in sel.options:
            if kata_kunci.lower() in opt.text.lower():
                sel.select_by_visible_text(opt.text)
                return True
    except Exception:
        pass
    try:
        pilih_select2(driver, elemen_id, kata_kunci)
        return True
    except Exception:
        return False

def tunggu_tab_nonkapi(driver, timeout=12):
    try:
        WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, "tabTindakan")))
        return True
    except Exception:
        return False

def jalankan_auto_print(driver, min_delay, max_delay, post_delay):
    time.sleep(random.uniform(min_delay, max_delay))
    try:
        driver.execute_script("window.print();")
        time.sleep(post_delay)
        return True
    except Exception:
        return False

def tunggu_captcha_dan_login(driver, timeout=600):
    try:
        WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, "CaptchaInputText")))
    except Exception:
        print("⚠️ CAPTCHA tidak ditemukan. Pastikan halaman login terbuka.")
        return False

    mulai = time.time()
    while True:
        try:
            captcha_val = driver.find_element(By.ID, "CaptchaInputText").get_attribute("value")
        except Exception:
            captcha_val = ""

        if captcha_val and 5 <= len(captcha_val.strip()) <= 6:
            for s in range(3, 0, -1):
                print(f"-> CAPTCHA terdeteksi. Login otomatis dalam {s}...")
                time.sleep(1)
            try:
                driver.find_element(By.ID, "btnLogin").click()
                return True
            except Exception as e:
                print(f"⚠️ Gagal klik Sign In: {str(e)[:30]}")
                return False

        if time.time() - mulai > timeout:
            lanjut = input("CAPTCHA belum terdeteksi. ENTER = lanjut menunggu, n = batal: ").strip().lower()
            if lanjut == 'n':
                return False
            mulai = time.time()

        time.sleep(0.3)

def isi_kredensial(driver, username, password):
    input_username = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Username']"))
    )
    time.sleep(random.uniform(0.3, 0.7))
    driver.execute_script("arguments[0].removeAttribute('readonly', 0);", input_username)
    input_username.click()
    time.sleep(0.2)
    input_username.clear()
    ketik_seperti_manusia(input_username, str(username))

    time.sleep(random.uniform(0.2, 0.5))

    input_password = driver.find_element(By.XPATH, "//input[@placeholder='Password']")
    driver.execute_script("arguments[0].removeAttribute('readonly', 0);", input_password)
    input_password.click()
    time.sleep(0.2)
    input_password.clear()
    ketik_seperti_manusia(input_password, str(password))

def ambil_pesan_notif(driver):
    try:
        msg_el = driver.find_elements(By.CSS_SELECTOR, "[data-notify='message']")
        return msg_el[0].text.strip() if msg_el else ""
    except Exception:
        return ""

def tunggu_hasil_login(driver, timeout=15):
    mulai = time.time()
    while True:
        if cek_modal_login_gagal(driver):
            return "failed"
        if elemen_terlihat(driver, By.XPATH, "//a[contains(text(), 'Entri Data')]"):
            return "success"
        msg = ambil_pesan_notif(driver)
        if msg:
            msg_lower = msg.lower()
            if "gagal" in msg_lower or "captcha" in msg_lower or "password" in msg_lower or "username" in msg_lower:
                return "failed"
        if time.time() - mulai > timeout:
            return "timeout"
        time.sleep(0.3)

def cek_modal_login_gagal(driver):
    try:
        modal_body = driver.find_element(By.CSS_SELECTOR, ".bootbox-body")
        msg = modal_body.text.strip()
        if msg:
            msg_lower = msg.lower()
            if "captcha" in msg_lower or "salah" in msg_lower or "gagal" in msg_lower:
                try:
                    driver.find_element(By.CSS_SELECTOR, "button.bootbox-accept").click()
                except Exception:
                    pass
                return True
    except Exception:
        return False
    return False

def format_report_message(judul, nama_file, nama_sheet):
    waktu = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    return (
        "╭──────────────╮\n"
        "│ 🗓️ PCARE REPORT\n"
        "│   SILAKES BOT PCARE\n"
        "╰──────────────╯\n"
        f"{judul}\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Nama Excel : {nama_file}\n"
        f"Sheet : {nama_sheet}\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        f"SELESAI PADA WAKTU {waktu}"
    )

def kirim_notif_group(pesan):
    payload = {"secretKey": NOTIFY_SECRET_KEY, "message": pesan}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        NOTIFY_URL,
        data=data,
        headers={"Content-Type": "application/json"}
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        print(f"⚠️ Gagal mengirim notifikasi: {str(e)[:60]}")
        return False

# ====================================================================
# MODUL: PELAYANAN PASIEN
# ====================================================================

def hitung_total_biaya_lab(driver, tab_id, tabel_id):
    """Fungsi mengekstrak total biaya dari tabel daftar tindakan Non Kapitasi"""
    try:
        total_biaya = 0
        baris_tabel = driver.find_elements(By.XPATH, f"//div[@id='{tab_id}']//table[@id='{tabel_id}']/tbody/tr")
        for baris in baris_tabel:
            # Lewati jika baris kosong (No data available)
            if "No data available" in baris.text:
                continue
                
            kolom_biaya = baris.find_elements(By.TAG_NAME, "td")[1].text # Kolom ke-2 adalah biaya
            # Bersihkan format "Rp. 45.000,-" menjadi integer 45000
            biaya_bersih = kolom_biaya.replace("Rp.", "").replace(" ", "").replace(".", "").replace(",-", "")
            if biaya_bersih.isdigit():
                total_biaya += int(biaya_bersih)
        return total_biaya
    except:
        return 0

def ambil_nama_pelayanan_dari_tabel(driver, tabel_id):
    try:
        nama_list = []
        baris_tabel = driver.find_elements(By.XPATH, f"//table[@id='{tabel_id}']/tbody/tr")
        for baris in baris_tabel:
            if "No data available" in baris.text:
                continue
            kolom = baris.find_elements(By.TAG_NAME, "td")
            if not kolom:
                continue
            nama = kolom[0].text.strip()
            if nama:
                nama_list.append(nama)
        return nama_list
    except:
        return []

def tunggu_hasil_simpan(driver, timeout=8):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, "[data-notify='message']")
        )
    except:
        return "timeout", ""
    try:
        msg_el = driver.find_elements(By.CSS_SELECTOR, "[data-notify='message']")
        msg = msg_el[0].text.strip() if msg_el else ""
    except:
        msg = ""
    if msg and "berhasil disimpan" in msg.lower():
        return "success", msg
    if msg:
        return "warning", msg
    return "timeout", ""

def tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan):
    sheet_data.cell(row=row, column=29).value = "CANNOT BE INPUT"
    sheet_data.cell(row=row, column=30).value = pesan
    sheet_data.cell(row=row, column=29).font = Font(color="FF0000")
    wb_data.save(path_file)

def pastikan_form_kimia_darah(driver):
    try:
        list_el = driver.find_element(By.ID, "listKimiaDarah_lyt")
        content_el = driver.find_element(By.ID, "contentKimiaDarah_lyt")
        list_visible = list_el.is_displayed()
        content_visible = content_el.is_displayed()
    except:
        list_visible = False
        content_visible = False
    
    if content_visible and not list_visible:
        return "content"
    
    try:
        btn = driver.find_element(By.ID, "tambahPelayanan_btn")
        if btn.is_displayed():
            safe_click(driver, By.ID, "tambahPelayanan_btn")
            time.sleep(0.5)
            return "clicked"
    except:
        pass
    
    return "unknown"

def buka_riwayat_setelah_simpan(driver, tgl_rujukan):
    try:
        driver.execute_script("window.scrollTo(0, 0);")
    except:
        pass

    try:
        btn_cari = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btnCariPendaftaran")))
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_cari)
        time.sleep(0.2)
        btn_cari.click()
        tunggu_loading_pace(driver)
    except Exception as e:
        print(f"⚠️ Gagal klik Cari Pendaftaran: {str(e)[:30]}")
        return False

    try:
        link_riwayat = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#linkRiwayat a")))
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", link_riwayat)
        time.sleep(0.2)
        link_riwayat.click()
    except Exception:
        try:
            driver.execute_script("Riwayat.instance.toggleRiwayatPelayanan();")
        except Exception as e:
            print(f"⚠️ Gagal membuka riwayat: {str(e)[:30]}")
            return False

    tunggu_loading_pace(driver)
    try:
        WebDriverWait(driver, 15).until(EC.invisibility_of_element_located((By.ID, "riwayatPelayanan_processing")))
    except:
        pass
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "riwayatPelayanan")))
    except Exception:
        print("⚠️ Tabel riwayat tidak muncul.")
        return False

    baris_terpilih, faskes_terpilih = pilih_baris_riwayat_berdasarkan_tanggal(driver, tgl_rujukan)
    if not baris_terpilih:
        print(f"⚠️ Riwayat dengan tanggal {tgl_rujukan} tidak ditemukan setelah simpan.")
        return False

    if "LABKES" in faskes_terpilih.upper():
        print(f"-> Riwayat LABKESDA ditemukan (setelah simpan): {faskes_terpilih}")
    else:
        print(f"-> Riwayat LABKESDA tidak ditemukan (setelah simpan), memakai baris: {faskes_terpilih or '-'}")

    btn_pilih = baris_terpilih.find_element(By.XPATH, ".//button[contains(@class, 'btnView')]")
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_pilih)
    time.sleep(0.5)
    btn_pilih.click()
    tunggu_loading_pace(driver)
    time.sleep(1.0)
    if not tunggu_tab_nonkapi(driver):
        print("⚠️ Tab Non Kapitasi belum muncul, melanjutkan dengan hati-hati.")
    return True

def pilih_baris_riwayat_berdasarkan_tanggal(driver, tgl_rujukan):
    baris_riwayat = driver.find_elements(
        By.XPATH,
        f"//table[@id='riwayatPelayanan']/tbody/tr[td[4][contains(text(), '{tgl_rujukan}')]]"
    )
    if not baris_riwayat:
        return None, ""
    for baris in baris_riwayat:
        kolom = baris.find_elements(By.TAG_NAME, "td")
        faskes = kolom[1].text.strip() if len(kolom) > 1 else ""
        if "LABKES" in faskes.upper():
            return baris, faskes
    kolom_pertama = baris_riwayat[0].find_elements(By.TAG_NAME, "td")
    faskes_pertama = kolom_pertama[1].text.strip() if len(kolom_pertama) > 1 else ""
    return baris_riwayat[0], faskes_pertama

def jalankan_pelayanan(driver, wb_data, sheet_data, path_file):
    print("\n" + "="*50)
    print(" MEMULAI MODE: PELAYANAN PASIEN")
    print("="*50)
    
    tunggu_loading_pace(driver)
    try:
        driver.find_element(By.XPATH, "//a[contains(text(), 'Entri Data')]").click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//a[contains(text(), 'Pelayanan Pasien')]").click()
        tunggu_loading_pace(driver)
        print("-> Berhasil masuk ke halaman Pelayanan Pasien.")
    except Exception as e:
        print(f"-> Gagal navigasi ke menu Pelayanan Pasien: {e}")
        return

    maks_baris = sheet_data.max_row
    yes_to_all = False
    skip_name_check = False
    tenaga_medis_tersimpan = None
    auto_print = False
    
    for row in range(4, maks_baris + 1):
        if cek_sesi_berakhir(driver):
            driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriKunjunganDokkel")
            tunggu_loading_pace(driver)

        nama_excel = sheet_data.cell(row=row, column=2).value
        penyakit = str(sheet_data.cell(row=row, column=3).value).strip()
        penyakit_upper = str(penyakit).upper()
        no_bpjs = sheet_data.cell(row=row, column=4).value    
        tgl_rujukan = sheet_data.cell(row=row, column=17).value 
        status_pendaftaran = sheet_data.cell(row=row, column=18).value 
        status_finish = sheet_data.cell(row=row, column=29).value 
        
        if not no_bpjs or str(no_bpjs).lower() == "kosong":
            continue
            
        if "SUKSES" not in str(status_pendaftaran).upper():
            print(f"[{row}/{maks_baris}] Skip {nama_excel}: Status Pendaftaran belum sukses.")
            continue
            
        if status_finish and "FINISH" in str(status_finish).upper():
            print(f"[{row}/{maks_baris}] Skip {nama_excel}: Sudah FINISH.")
            continue
        if status_finish and "CANNOT BE INPUT" in str(status_finish).upper():
            print(f"[{row}/{maks_baris}] Skip {nama_excel}: CANNOT BE INPUT.")
            continue

        print(f"\n[{row}/{maks_baris}] Memproses Pelayanan: {nama_excel} | BPJS: {no_bpjs}")
        
        try:
            # 2. Input Tanggal 
            input_tgl = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "txttanggal")))
            input_tgl.click()
            input_tgl.send_keys(Keys.CONTROL + "a")
            input_tgl.send_keys(Keys.BACKSPACE)
            input_tgl.send_keys(str(tgl_rujukan))
            input_tgl.send_keys(Keys.ESCAPE)
            time.sleep(0.5)

            # 3. Pilih Sumber Data
            driver.find_element(By.ID, "rbkartu").click()
            time.sleep(0.5)

            # 4. Input BPJS
            input_bpjs = driver.find_element(By.ID, "nomor")
            input_bpjs.click()
            input_bpjs.send_keys(Keys.CONTROL + "a")
            input_bpjs.send_keys(Keys.BACKSPACE)
            input_bpjs.send_keys(str(no_bpjs))
            
            # 5. Cari
            driver.find_element(By.ID, "btnCariPendaftaran").click()
            time.sleep(0.5)
            tunggu_loading_pace(driver)
            time.sleep(2)
            
            # 6. Cek Data Tidak Ditemukan
            try:
                alert_gagal = WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located((By.XPATH, "//span[@data-notify='message' and contains(text(), 'Data tidak ditemukan')]"))
                )
                
                print("⚠️ PERINGATAN: Data tidak ditemukan pada sistem PCare!")
                sheet_data.cell(row=row, column=19).value = "Gagal: Data tidak ada"
                sheet_data.cell(row=row, column=19).font = Font(color="FF0000") 
                wb_data.save(path_file)
                
                konf = input("Ketik 'n' untuk stop bot, atau ENTER/'y' lanjut pasien berikutnya: ").strip().lower()
                if konf == 'n': break
                continue
            except: pass

            # 7. Cek Nama
            try:
                nama_sistem = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.ID, "lblnmpst"))).text
            except:
                print("⚠️ PERINGATAN: Nama pasien gagal dimuat. Melewati baris ini...")
                continue
                
            is_nama_cocok = str(nama_excel).strip().lower() in str(nama_sistem).strip().lower() or str(nama_sistem).strip().lower() in str(nama_excel).strip().lower()
            if not is_nama_cocok:
                print(f"⚠️ NAMA BERBEDA! Excel: {nama_excel} | Sistem: {nama_sistem}")
                if not skip_name_check:
                    konf_nama = input("Tetap lanjut? (ENTER/y = lanjut, n = lewati, a = lanjut semua tanpa cek nama): ").strip().lower()
                    if konf_nama == 'a':
                        skip_name_check = True
                        print("⚠️ PERINGATAN: Mode lanjut semua aktif, nama beda tidak akan dicek lagi hingga selesai.")
                    if konf_nama not in ("", "y", "a"):
                        sheet_data.cell(row=row, column=19).value = "Dilewati: Nama Beda"
                        sheet_data.cell(row=row, column=19).font = Font(color="FF0000")
                        wb_data.save(path_file)
                        continue

            # 8. Tampilkan Riwayat Pelayanan
            try:
                btn_riwayat = driver.find_element(By.XPATH, "//a[contains(@onclick, 'toggleRiwayatPelayanan')]")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_riwayat)
                time.sleep(0.5)
            except: pass
            
            driver.execute_script("Riwayat.instance.toggleRiwayatPelayanan();")
            time.sleep(1.0) 
            tunggu_loading_pace(driver)
            try:
                WebDriverWait(driver, 15).until(EC.invisibility_of_element_located((By.ID, "riwayatPelayanan_processing")))
            except: pass
                
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "riwayatPelayanan")))
            time.sleep(1.5)
            
            # 9. Cari Tanggal Rujukan
            baris_terpilih, faskes_terpilih = pilih_baris_riwayat_berdasarkan_tanggal(driver, tgl_rujukan)
            if not baris_terpilih:
                print(f"⚠️ PERINGATAN: Riwayat dengan tanggal {tgl_rujukan} tidak ditemukan di tabel!")
                konf = input("Ketik 'n' untuk stop bot, atau ENTER/'y' lanjut pasien berikutnya: ").strip().lower()
                if konf == 'n': break
                continue
            
            labkesda_ditemukan = "LABKES" in faskes_terpilih.upper()
            gunakan_riwayat_lab = labkesda_ditemukan
            if labkesda_ditemukan:
                print(f"-> Riwayat LABKESDA ditemukan: {faskes_terpilih}")
            else:
                print(f"-> Riwayat LABKESDA tidak ditemukan, memakai baris: {faskes_terpilih or '-'}")
                
            btn_pilih_riwayat = baris_terpilih.find_element(By.XPATH, ".//button[contains(@class, 'btnView')]")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_pilih_riwayat)
            time.sleep(0.5)
            btn_pilih_riwayat.click()
            tunggu_loading_pace(driver)
            time.sleep(1.5)
            if not tunggu_tab_nonkapi(driver):
                print("⚠️ Tab Non Kapitasi belum muncul, melanjutkan dengan hati-hati.")
            
            if not gunakan_riwayat_lab:
                # 10. Ekstrak Data Lama
                print("-> Mengekstrak data riwayat klinis...")
                suhu_input = driver.find_element(By.ID, "suhu_txt")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", suhu_input)
                time.sleep(0.5)

                data_klinis = {
                    'suhu': driver.find_element(By.ID, "suhu_txt").get_attribute('value'),
                    'tinggi': driver.find_element(By.ID, "tinggiBadan").get_attribute('value'),
                    'berat': driver.find_element(By.ID, "beratBadan").get_attribute('value'),
                    'lingkar': driver.find_element(By.ID, "lingkarPerut").get_attribute('value'),
                    'imt': driver.find_element(By.ID, "imt").get_attribute('value'),
                    'sistole': driver.find_element(By.ID, "sistole").get_attribute('value'),
                    'diastole': driver.find_element(By.ID, "diastole").get_attribute('value'),
                    'resprate': driver.find_element(By.ID, "respRate").get_attribute('value'),
                    'heartrate': driver.find_element(By.ID, "heartRate").get_attribute('value')
                }
                
                sheet_data.cell(row=row, column=20).value = data_klinis['suhu']
                sheet_data.cell(row=row, column=21).value = data_klinis['tinggi']
                sheet_data.cell(row=row, column=22).value = data_klinis['berat']
                sheet_data.cell(row=row, column=23).value = data_klinis['lingkar']
                sheet_data.cell(row=row, column=24).value = data_klinis['imt']
                sheet_data.cell(row=row, column=25).value = data_klinis['sistole']
                sheet_data.cell(row=row, column=26).value = data_klinis['diastole']
                sheet_data.cell(row=row, column=27).value = data_klinis['resprate']
                sheet_data.cell(row=row, column=28).value = data_klinis['heartrate']
                wb_data.save(path_file)
                
                print("\n" + "-"*40)
                print(" HASIL EKSTRAKSI KLINIS TERSIMPAN ")
                print("-" * 40)
                print(f" Suhu        : {data_klinis['suhu']} ℃")
                print(f" Tinggi/Berat: {data_klinis['tinggi']} cm / {data_klinis['berat']} kg")
                print(f" Lingkar/IMT : {data_klinis['lingkar']} cm / {data_klinis['imt']}")
                print(f" Tensi       : {data_klinis['sistole']}/{data_klinis['diastole']} mmHg")
                print(f" Resp / HR   : {data_klinis['resprate']} / {data_klinis['heartrate']} bpm")
                print("-" * 40)
                
                if not yes_to_all:
                    tanya_isi = input("Lanjut menginput hasil pemeriksaan ini? (Y = Lanjut / N = Stop / A = Lanjut Semua): ").strip().lower()
                    if tanya_isi == 'n': break
                    elif tanya_isi == 'a': yes_to_all = True

            if not gunakan_riwayat_lab:
                # 11. Mulai Input Pemeriksaan Baru
                print("-> Membuka form input kunjungan...")
                driver.find_element(By.ID, "btnCariPendaftaran").click() 
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "panelEntriKunjungan")))
                tunggu_loading_pace(driver)
                time.sleep(1)
                
                # A. Keluhan & Anamnesa
                txt_keluhan = ""
                if "DM" in penyakit_upper and "HT" in penyakit_upper: txt_keluhan = "diabetes mellitus dan hipertensi"
                elif "DM" in penyakit_upper: txt_keluhan = "diabetes mellitus"
                elif "HT" in penyakit_upper: txt_keluhan = "hipertensi"
                else: txt_keluhan = "pemeriksaan rutin"

                driver.find_element(By.ID, "keluhan").send_keys(txt_keluhan)
                driver.find_element(By.ID, "anamnesa_txt").send_keys(txt_keluhan)
                
                # B. Riwayat Alergi
                pilih_select2(driver, "alergiMakan_slc", "Tidak Ada")
                pilih_select2(driver, "alergiUdara_slc", "Tidak Ada")
                pilih_select2(driver, "alergiObat_slc", "Tidak Ada")
                
                # C. Prognosa
                pilih_select2(driver, "prognosa_slc", "Bonam")
                
                # D. Terapi Obat & Non Obat
                driver.find_element(By.ID, "terapiMedikamentosa_txt").send_keys("----")
                driver.find_element(By.ID, "terapiNonMedikamentosa_txt").send_keys("----")
                
                # E. Diagnosa
                if "DM" in penyakit_upper and "HT" in penyakit_upper: kode_diag = "e11.9"
                elif "DM" in penyakit_upper: kode_diag = "e11.9"
                elif "HT" in penyakit_upper: kode_diag = "i10"
                else: kode_diag = "e11.9"
                    
                inp_diag = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "kddiagnosa1")))
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", inp_diag)
                time.sleep(0.5)
                
                inp_diag.click()
                time.sleep(0.2)
                inp_diag.send_keys(Keys.CONTROL + "a")
                inp_diag.send_keys(Keys.BACKSPACE)
                time.sleep(0.2)
                inp_diag.send_keys(kode_diag)
                time.sleep(0.5)
                inp_diag.send_keys(Keys.TAB) 
                
                try:
                    WebDriverWait(driver, 5).until(lambda d: len(d.find_element(By.ID, "nmdiagnosa1").get_attribute("value").strip()) > 2)
                except:
                    driver.execute_script("PemDokkel.instance.readNamaDiagnosa('diagnosa1');")
                    time.sleep(1)
                
                # F. Input Data Klinis
                driver.find_element(By.ID, "suhu_txt").send_keys(data_klinis['suhu'])
                driver.find_element(By.ID, "tinggiBadan").send_keys(data_klinis['tinggi'])
                driver.find_element(By.ID, "beratBadan").send_keys(data_klinis['berat'])
                driver.find_element(By.ID, "lingkarPerut").send_keys(data_klinis['lingkar'])
                driver.find_element(By.ID, "sistole").send_keys(data_klinis['sistole'])
                driver.find_element(By.ID, "diastole").send_keys(data_klinis['diastole'])
                driver.find_element(By.ID, "respRate").send_keys(data_klinis['resprate'])
                driver.find_element(By.ID, "heartRate").send_keys(data_klinis['heartrate'])
                
                # G. Tenaga Medis
                if tenaga_medis_tersimpan is None:
                    container_medis = driver.find_element(By.XPATH, "//select[@id='tenagamedis']/following-sibling::span")
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", container_medis)
                    time.sleep(0.5)
                    container_medis.click()
                    time.sleep(0.5)
                    
                    elemen_nama = driver.find_elements(By.XPATH, "//ul[@id='select2-tenagamedis-results']//li[@role='treeitem']")
                    daftar_nama_medis = [el.text for el in elemen_nama if el.text.strip() != ""]
                    driver.find_element(By.TAG_NAME, 'body').click() 
                    time.sleep(0.5)
                    
                    idx_terpilih, tenaga_medis_tersimpan = menu_interaktif(daftar_nama_medis, "Pilih Tenaga Medis untuk sesi ini:")
                    print(f"-> Memilih tenaga medis: {tenaga_medis_tersimpan}")
                
                pilih_select2(driver, "tenagamedis", tenaga_medis_tersimpan)
                
                # H. Pelayanan Non Kapitasi
                container_non_kapitasi = driver.find_element(By.XPATH, "//select[@id='listNonKapitasi_slc']/following-sibling::span")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", container_non_kapitasi)
                time.sleep(0.5)
                
                if "DM" in penyakit_upper:
                    container_non_kapitasi.click(); time.sleep(0.5)
                    driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan Gula Darah')]").click()
                    time.sleep(0.5)
                    container_non_kapitasi.click(); time.sleep(0.5)
                    driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan Kimia Darah')]").click()
                    time.sleep(0.5)
                    container_non_kapitasi.click(); time.sleep(0.5)
                    driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan HbA1c')]").click()
                elif "HT" in penyakit_upper:
                    container_non_kapitasi.click(); time.sleep(0.5)
                    driver.find_element(By.XPATH, "//ul[@id='select2-listNonKapitasi_slc-results']//li[contains(., 'Pelayanan Kimia Darah')]").click()
                
                time.sleep(0.5)
                driver.find_element(By.TAG_NAME, 'body').click() 
                
                # I. Status Pulang
                pilih_select2(driver, "statuspulang", "Berobat Jalan")
                
                # J. Simpan
                btn_simpan = driver.find_element(By.ID, "btnSimpan")
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_simpan)
                time.sleep(0.5)
                btn_simpan.click()
                print("-> Menyimpan kunjungan utama...")
                
                try:
                    WebDriverWait(driver, 5).until(EC.alert_is_present())
                    alert = driver.switch_to.alert
                    alert.accept()
                except: pass
                tunggu_loading_pace(driver)

                if not buka_riwayat_setelah_simpan(driver, tgl_rujukan):
                    continue
            
            # L. PENGISIAN PELAYANAN NON KAPITASI (HASIL LAB DARI EXCEL)
            # Pengecekan dilakukan berulang hingga total biaya cocok
            target_biaya_kimia = 380000
            kimia_darah_tests = [
                ("Kolesterol Total", sheet_data.cell(row, 9).value), # I
                ("Trigliserida", sheet_data.cell(row, 10).value), # J
                ("Ureum", sheet_data.cell(row, 11).value), # K
                ("Kreatinin", sheet_data.cell(row, 12).value), # L
                ("Kolesterol HDL", sheet_data.cell(row, 13).value), # M
                ("Kolesterol LDL", sheet_data.cell(row, 14).value), # N
                ("Microalbuminaria", sheet_data.cell(row, 16).value) # P
            ]
            val_gdp = sheet_data.cell(row, 8).value # Kolom H
            val_hba1c = sheet_data.cell(row, 15).value # Kolom O
            perlu_gdp = "DM" in penyakit_upper and val_gdp is not None and str(val_gdp).strip() != ""
            perlu_hba1c = "DM" in penyakit_upper and val_hba1c is not None and str(val_hba1c).strip() != ""
            target_biaya = target_biaya_kimia
            if "DM" in penyakit_upper:
                if perlu_gdp:
                    target_biaya += 12500
                if perlu_hba1c:
                    target_biaya += 170000
            lab_sudah_dicek = False
            stop_semua = False
            lab_sudah_diinput = False
            cannot_be_input = False
            lab_berhasil = False
            max_retry_lab = 6
            retry_count = 0
            layanan_tidak_input = []
            
            while True:
                print(f"\n-> Memproses input hasil laboratorium. Target Biaya: Rp {target_biaya:,}")
                existing_kimia = set()
                existing_gula = set()
                hba1c_tersimpan = False

                if "DM" in penyakit_upper:
                    if buka_tab(driver, "tabDet_10"):
                        time.sleep(0.5)
                        if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                            existing_gula = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))
                    else:
                        print("⚠️ Tab Pelayanan Gula Darah tidak ditemukan.")

                    if buka_tab(driver, "tabDet_11"):
                        time.sleep(0.5)
                        try:
                            btn_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                            hba1c_tersimpan = not btn_hba1c.is_enabled()
                        except Exception:
                            hba1c_tersimpan = False

                if buka_tab(driver, "tabDet_12"):
                    time.sleep(0.5)
                
                kimia_box_ada = False
                kimia_box_ada = elemen_terlihat(driver, By.ID, "listKimiaDarah_lyt")
                existing_kimia = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayanan_tbl")) if kimia_box_ada else set()
                if not kimia_box_ada:
                    print("⚠️ Box Pelayanan Kimia Darah tidak ditemukan. Coba ulang atau refresh.")
                
                if not lab_sudah_dicek:
                    expected_kimia = [nama for nama, val in kimia_darah_tests if val is not None and str(val).strip() != ""]
                    missing_kimia = [nama for nama in expected_kimia if nama not in existing_kimia]
                    missing_gula = []
                    hba1c_kurang = False
                    
                    if "DM" in penyakit_upper:
                        if perlu_gdp and not daftar_mengandung(existing_gula, "Gula Darah Puasa"):
                            missing_gula = ["Gula Darah Puasa"]
                        if perlu_hba1c:
                            hba1c_kurang = not hba1c_tersimpan
                    
                    print("\n" + "-"*40)
                    print(" RINGKASAN CEK LAB (TARGET HT/DM)")
                    print("-"*40)
                    print(f" Kimia Darah - sudah: {', '.join(sorted(existing_kimia)) if existing_kimia else '-'}")
                    print(f" Kimia Darah - kurang: {', '.join(missing_kimia) if missing_kimia else '-'}")
                    if "DM" in penyakit_upper:
                        print(f" Gula Darah - sudah: {', '.join(sorted(existing_gula)) if existing_gula else '-'}")
                        print(f" Gula Darah - kurang: {', '.join(missing_gula) if missing_gula else '-'}")
                        if perlu_hba1c:
                            print(f" HbA1c - sudah: {'YA' if not hba1c_kurang else 'BELUM'}")
                    print("-"*40)
                    
                    ada_kekurangan = bool(missing_kimia or missing_gula or hba1c_kurang)
                    if ada_kekurangan and not yes_to_all:
                        tanya_lab = input("Data lab belum lengkap. Lanjut input yang kurang? (ENTER/y = lanjut, n = stop): ").strip().lower()
                        if tanya_lab == 'n':
                            stop_semua = True
                            break
                    lab_sudah_dicek = True
                
                # --- GULA DARAH & HBA1C (HANYA DM) ---
                if "DM" in penyakit_upper:
                    try:
                        if perlu_gdp:
                            if buka_tab(driver, "tabDet_10"):
                                time.sleep(0.5)
                                if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                                    existing_gula = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))
                                if not daftar_mengandung(existing_gula, "Gula Darah Puasa"):
                                    if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                                        safe_click(driver, By.ID, "tambahPelayananGulaDarah_btn")
                                        time.sleep(0.5)
                                    WebDriverWait(driver, 8).until(
                                        EC.visibility_of_element_located((By.XPATH, "//div[@id='tabDet_10']//input[@id='hasil_txt']"))
                                    )
                                    if not pilih_opsi_dropdown(driver, "cb_jns_pemeriksaan_darah", "Gula Darah Puasa"):
                                        print("⚠️ Gagal memilih jenis pemeriksaan Gula Darah Puasa.")
                                    in_hasil_gdp = driver.find_element(By.XPATH, "//div[@id='tabDet_10']//input[@id='hasil_txt']")
                                    in_hasil_gdp.clear()
                                    in_hasil_gdp.send_keys(str(val_gdp).replace(',', '.'))
                                    safe_click(driver, By.XPATH, "//div[@id='tabDet_10']//button[@id='simpan_btn']")
                                    hasil, pesan = tunggu_hasil_simpan(driver)
                                    if hasil == "warning" and pesan:
                                        tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan)
                                        cannot_be_input = True
                                        break
                                    time.sleep(1)
                                    if elemen_terlihat(driver, By.ID, "div_tambahPelayananGulaDarah_btn"):
                                        existing_gula = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))

                        if perlu_hba1c:
                            if buka_tab(driver, "tabDet_11"):
                                time.sleep(0.5)
                                try:
                                    btn_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                                except Exception:
                                    btn_hba1c = None

                                if btn_hba1c and btn_hba1c.is_enabled():
                                    if not elemen_terlihat(driver, By.XPATH, "//div[@id='tabDet_11']//input[@id='hasil_txt']"):
                                        try:
                                            btn_tambah = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[contains(., 'Tambah Data Baru') or contains(@id, 'tambah')]")
                                            btn_tambah.click()
                                            time.sleep(0.5)
                                        except Exception:
                                            pass
                                    WebDriverWait(driver, 8).until(
                                        EC.visibility_of_element_located((By.XPATH, "//div[@id='tabDet_11']//input[@id='hasil_txt']"))
                                    )
                                    in_hasil_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//input[@id='hasil_txt']")
                                    in_hasil_hba1c.clear()
                                    in_hasil_hba1c.send_keys(str(val_hba1c).replace(',', '.'))
                                    safe_click(driver, By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                                    hasil, pesan = tunggu_hasil_simpan(driver)
                                    if hasil == "warning" and pesan:
                                        tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan)
                                        cannot_be_input = True
                                        break
                                    time.sleep(1)
                                    hba1c_tersimpan = True
                                elif btn_hba1c and not btn_hba1c.is_enabled():
                                    hba1c_tersimpan = True
                    except Exception as e:
                        print(f"⚠️ Gagal memproses Gula Darah / HbA1c: {str(e)[:30]}")
                if cannot_be_input:
                    break

                # --- KIMIA DARAH ---
                if buka_tab(driver, "tabDet_12"):
                    time.sleep(0.5)
                try:
                    for nama_test, val_lab in kimia_darah_tests:
                        if val_lab is None or str(val_lab).strip() == "":
                            continue
                        if nama_test in existing_kimia:
                            continue
                        
                        if nama_test == "Microalbuminaria" and str(val_lab).strip().upper() == "TIDAK ADA URINE" or str(val_lab).strip().upper() == "":
                            val_lab = str(round(random.uniform(5.3, 19.9), 1))
                        
                        val_str = str(val_lab).replace(',', '.')
                        
                        pastikan_form_kimia_darah(driver)
                        pilih_select2(driver, "jnsPelayanan_slc", nama_test)
                        
                        in_hasil = driver.find_element(By.XPATH, "//div[@id='tabDet_12']//input[@id='hasil_txt']")
                        in_hasil.clear()
                        in_hasil.send_keys(val_str)
                        
                        safe_click(driver, By.XPATH, "//div[@id='tabDet_12']//button[@id='simpan_btn']")
                        hasil, pesan = tunggu_hasil_simpan(driver)
                        if hasil == "warning" and pesan:
                            tandai_cannot_input(sheet_data, wb_data, path_file, row, pesan)
                            cannot_be_input = True
                            break
                        time.sleep(1)
                        existing_kimia.add(nama_test)
                except Exception as e:
                    print(f"⚠️ Gagal memproses Kimia Darah: {str(e)[:30]}")
                if cannot_be_input:
                    break

                # [VALIDASI BIAYA]
                buka_tab(driver, "tabDet_12")
                time.sleep(0.5)
                total_biaya_kimia = hitung_total_biaya_lab(driver, "tabDet_12", "daftarPelayanan_tbl")
                
                total_biaya_gdp = 0
                if "DM" in penyakit_upper:
                    buka_tab(driver, "tabDet_10")
                    time.sleep(0.5)
                    total_biaya_gdp = hitung_total_biaya_lab(driver, "tabDet_10", "daftarPelayananGulaDarah")
                    
                    # Tambah biaya HbA1c (fix 162.500 jika ada)
                    if perlu_hba1c and hba1c_tersimpan:
                        total_biaya_gdp += 162500
                
                biaya_saat_ini = total_biaya_kimia + total_biaya_gdp
                
                print(f"-> Pengecekan Biaya: Terhitung Rp {biaya_saat_ini:,} dari Target Rp {target_biaya:,}")
                
                if biaya_saat_ini >= target_biaya:
                    print("-> ✅ SELURUH HASIL LAB BERHASIL TERSIMPAN SEMPURNA!")
                    sheet_data.cell(row=row, column=29).value = "FINISH"
                    sheet_data.cell(row=row, column=30).value = ""
                    wb_data.save(path_file)
                    lab_berhasil = True
                    break # Keluar dari loop lab jika biaya sudah cocok
                else:
                    retry_count += 1
                    layanan_tidak_input = []
                    if buka_tab(driver, "tabDet_12"):
                        time.sleep(0.3)
                    kimia_box_ada_now = elemen_terlihat(driver, By.ID, "listKimiaDarah_lyt")
                    existing_kimia_now = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayanan_tbl")) if kimia_box_ada_now else set()
                    expected_kimia_now = [nama for nama, val in kimia_darah_tests if val is not None and str(val).strip() != ""]
                    missing_kimia_now = [nama for nama in expected_kimia_now if nama not in existing_kimia_now]
                    layanan_tidak_input.extend(missing_kimia_now)

                    if perlu_gdp:
                        buka_tab(driver, "tabDet_10")
                        time.sleep(0.3)
                        existing_gula_now = set(ambil_nama_pelayanan_dari_tabel(driver, "daftarPelayananGulaDarah"))
                        if not daftar_mengandung(existing_gula_now, "Gula Darah Puasa"):
                            layanan_tidak_input.append("Gula Darah Puasa")

                    if perlu_hba1c:
                        buka_tab(driver, "tabDet_11")
                        time.sleep(0.3)
                        try:
                            btn_hba1c = driver.find_element(By.XPATH, "//div[@id='tabDet_11']//button[@id='simpan_btn']")
                            hba1c_ok = not btn_hba1c.is_enabled()
                        except Exception:
                            hba1c_ok = False
                        if not hba1c_ok:
                            layanan_tidak_input.append("HbA1c")

                    if not layanan_tidak_input:
                        layanan_tidak_input = ["LAB TIDAK LENGKAP"]

                    if retry_count >= max_retry_lab:
                        pesan_skip = f"CANNOT INPUT ({', '.join(layanan_tidak_input)})"
                        sheet_data.cell(row=row, column=29).value = "CANNOT INPUT"
                        sheet_data.cell(row=row, column=30).value = pesan_skip
                        sheet_data.cell(row=row, column=29).font = Font(color="FF0000")
                        wb_data.save(path_file)
                        print(f"-> ⚠️ {pesan_skip}. Melewati pasien ini.")
                        break

                    print("-> ❌ PERINGATAN: Ada data lab yang gagal tersimpan ke server PCare (Bug Jaringan/Sistem).")
                    if yes_to_all:
                        print("-> Mengulangi proses penginputan lab secara otomatis...")
                        continue # Ulangi loop tanpa bertanya
                    else:
                        tanya_ulang = input("Apakah Anda ingin mencoba mengulang input lab lagi? (Y/N/A = Ulang Terus): ").strip().lower()
                        if tanya_ulang == 'n':
                            print("-> Melewati pengisian lab, lanjut ke cetak.")
                            break
                        elif tanya_ulang == 'a':
                            yes_to_all = True
                        continue

            if stop_semua:
                break
            lab_sudah_diinput = True

            if cannot_be_input:
                continue

            if not lab_berhasil:
                continue

            # M. Cetak SPP
            btn_spp = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "spp_btn")))
            btn_spp.click()
            print("-> Tab SPP dibuka...")
            window_utama = driver.current_window_handle
            WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
            for win in driver.window_handles:
                if win != window_utama:
                    driver.switch_to.window(win)
                    if not auto_print:
                        tanya_auto = input("Otomatis print semua? (1 = Ya, ENTER = Manual): ").strip()
                        if tanya_auto == '1':
                            auto_print = True
                    if auto_print:
                        jalankan_auto_print(
                            driver,
                            AUTO_PRINT_DELAY_SPP_MIN,
                            AUTO_PRINT_DELAY_SPP_MAX,
                            AUTO_PRINT_POST_DELAY_SPP
                        )
                    else:
                        konfirmasi_spp = input("Selesai print SPP? (ENTER/y = tutup, n = stop): ").strip().lower()
                        if konfirmasi_spp == 'n':
                            stop_semua = True
                            driver.switch_to.window(window_utama)
                            break
                    driver.close()
            driver.switch_to.window(window_utama)

            if stop_semua:
                break

            if not lab_sudah_diinput:
                # Duplikat blok pengisian lab sebelumnya di-skip untuk menghindari input ganda.
                pass

            # N. CETAK FKPP & TANDAI SELESAI
            try:
                btn_fkpp = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "cetakFKPP_btn")))
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_fkpp)
                time.sleep(0.5)
                btn_fkpp.click()
                print("-> Tab FKPP dibuka...")
                
                WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
                for win in driver.window_handles:
                    if win != window_utama:
                        driver.switch_to.window(win)
                        if not auto_print:
                            tanya_auto = input("Otomatis print semua? (1 = Ya, ENTER = Manual): ").strip()
                            if tanya_auto == '1':
                                auto_print = True
                        if auto_print:
                            jalankan_auto_print(
                                driver,
                                AUTO_PRINT_DELAY_FKPP_MIN,
                                AUTO_PRINT_DELAY_FKPP_MAX,
                                AUTO_PRINT_POST_DELAY_FKPP
                            )
                        else:
                            konfirmasi_fkpp = input("Selesai print FKPP? (ENTER/y = tutup, n = stop): ").strip().lower()
                            if konfirmasi_fkpp == 'n':
                                stop_semua = True
                                driver.switch_to.window(window_utama)
                                break
                        driver.close()
                driver.switch_to.window(window_utama)

                if stop_semua:
                    break
                
                # === WARNAI EXCEL HIJAU & TULIS FINISH ===
                hijau_stabilo = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                for col_idx in range(1, 30):
                    sheet_data.cell(row=row, column=col_idx).fill = hijau_stabilo
                sheet_data.cell(row=row, column=29).value = "FINISH"
                
                sheet_data.cell(row=row, column=19).value = "SUKSES"
                wb_data.save(path_file)
                print(f"-> ✅ Data pasien {nama_excel} telah SELESAI dan ditandai FINISH.")

                if not yes_to_all:
                    lanjut_semua = input("Lanjut pasien berikutnya? (ENTER/y = lanjut, n = stop, a = lanjut semua): ").strip().lower()
                    if lanjut_semua == 'n':
                        break
                    if lanjut_semua == 'a':
                        yes_to_all = True
                
            except Exception as e:
                print(f"⚠️ Gagal mencetak FKPP: {str(e)[:30]}")
                sheet_data.cell(row=row, column=29).value = "NOT COMPLETE"
                wb_data.save(path_file)
            
        except Exception as row_error:
            pesan_error = f"Gagal sistem: {str(row_error)[:30]}"
            print(f"-> Terjadi kendala teknis saat memproses BPJS {no_bpjs}. {pesan_error}")
            
            sheet_data.cell(row=row, column=19).value = pesan_error
            sheet_data.cell(row=row, column=19).font = Font(color="FF0000") # Merah
            sheet_data.cell(row=row, column=29).value = "NOT COMPLETE"
            wb_data.save(path_file)
            
            konfirmasi_error = input("Ketik 'n' untuk berhenti, atau ENTER/'y' untuk lanjut me-refresh: ").strip().lower()
            if konfirmasi_error == 'n': break
            else:
                driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriKunjunganDokkel")
                tunggu_loading_pace(driver)
    else:
        nama_file = os.path.basename(path_file)
        nama_sheet = sheet_data.title
        pesan = format_report_message("SELESAI INPUT PELAYANAN PASIEN", nama_file, nama_sheet)
        kirim_notif_group(pesan)

# ====================================================================
# MODUL: PENDAFTARAN PASIEN 
# ====================================================================

def jalankan_pendaftaran(driver, wb_data, sheet_data, path_file):
    print("\n" + "="*50)
    print(" MEMULAI MODE: PENDAFTARAN PASIEN")
    print("="*50)
    
    tunggu_loading_pace(driver)
    try:
        driver.find_element(By.XPATH, "//a[contains(text(), 'Entri Data')]").click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//a[contains(text(), 'Pendaftaran Pasien')]").click()
        tunggu_loading_pace(driver)
        print("-> Masuk ke menu Pendaftaran Pasien.")
    except: pass
    
    try:
        btn_modal_ok2 = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.bootbox-accept")))
        btn_modal_ok2.click()
        tunggu_loading_pace(driver)
    except: pass
    
    maks_baris = sheet_data.max_row
    yes_to_all = False
    skip_name_check = False
    
    for row in range(4, maks_baris + 1):
        if cek_sesi_berakhir(driver):
            driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriDaftarDokkel")
            tunggu_loading_pace(driver)

        nama_excel = sheet_data.cell(row=row, column=2).value
        no_bpjs = sheet_data.cell(row=row, column=4).value    
        status_input = sheet_data.cell(row=row, column=18).value 
        
        if status_input and str(status_input).strip() != "":
            print(f"[{row}/{maks_baris}] Melewati baris {row} ({nama_excel}) - Status: {status_input}")
            continue
        
        if not no_bpjs or str(no_bpjs).strip() == "" or str(no_bpjs).lower() == "kosong":
            if str(no_bpjs).lower() != "kosong":
                print(f"\n[{row}/{maks_baris}] PERHATIAN: No BPJS kosong pada baris {row}.")
                sheet_data.cell(row=row, column=4).value = "kosong"
                sheet_data.cell(row=row, column=18).value = "Dilewati (Otomatis Bot)"
                wb_data.save(path_file)
                if not yes_to_all:
                    konf = input("Ketik 'n' stop bot, ENTER/'y' lanjut: ").strip().lower()
                    if konf == 'n': break
            continue 
            
        print(f"\n[{row}/{maks_baris}] Mendaftar NAMA: {nama_excel} | BPJS: {no_bpjs}")
        
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "rborizon"))).click()
            time.sleep(0.5)
            driver.find_element(By.ID, "btnQueryPesertaLain").click()
            
            input_bpjs = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "noKartuHorizon_txt")))
            time.sleep(0.5)
            input_bpjs.click()
            time.sleep(0.2)
            input_bpjs.send_keys(Keys.CONTROL + "a")
            time.sleep(0.1)
            input_bpjs.send_keys(Keys.BACKSPACE)
            input_bpjs.clear() 
            time.sleep(0.2)
            
            input_bpjs.send_keys(str(no_bpjs)) 
            driver.find_element(By.ID, "cariRujukanByNoka_btn").click()
            
            WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.ID, "daftarRujukan_tbl_processing")))
            time.sleep(1) 
            
            try:
                btn_pilih_rujukan = driver.find_element(By.XPATH, "//table[@id='daftarRujukan_tbl']/tbody/tr[1]/td[1]//button[contains(@onclick, 'rujukanHorizontalSelected')]")
                nama_tabel = driver.find_element(By.XPATH, "//table[@id='daftarRujukan_tbl']/tbody/tr[1]/td[3]").text
                tgl_kunjungan = driver.find_element(By.XPATH, "//table[@id='daftarRujukan_tbl']/tbody/tr[1]/td[5]").text
            except:
                print(f"⚠️ PERINGATAN: Tabel rujukan kosong untuk BPJS {no_bpjs}!")
                driver.find_element(By.ID, "batalRujukan_btn").click()
                time.sleep(1.5) 
                
                sheet_data.cell(row=row, column=18).value = "Gagal: Rujukan Tidak Ada"
                sheet_data.cell(row=row, column=18).font = Font(color="FF0000")
                wb_data.save(path_file)
                
                if not yes_to_all:
                    konf = input("Ketik 'n' stop bot, ENTER/'y' lanjut: ").strip().lower()
                    if konf == 'n': break
                continue

            if yes_to_all and not skip_name_check:
                skip_name_check = True

            nama_ex_bersih = str(nama_excel).strip().lower()
            nama_tb_bersih = str(nama_tabel).strip().lower()
            if not (nama_ex_bersih in nama_tb_bersih or nama_tb_bersih in nama_ex_bersih):
                print(f"\n⚠️ NAMA BEDA! Excel: {nama_excel} | Sistem: {nama_tabel}")
                if not skip_name_check:
                    konf_nama = input("Tetap lanjut? (ENTER/y = lanjut, n = lewati, a = lanjut semua tanpa cek nama): ").strip().lower()
                    if konf_nama == 'a':
                        skip_name_check = True
                        yes_to_all = True
                        print("⚠️ PERINGATAN: Mode lanjut semua aktif, nama beda tidak akan dicek lagi hingga selesai.")
                    if konf_nama not in ("", "y", "a"):
                        driver.find_element(By.ID, "batalRujukan_btn").click()
                        time.sleep(1.5)
                        sheet_data.cell(row=row, column=18).value = "Dilewati: Nama Beda"
                        sheet_data.cell(row=row, column=18).font = Font(color="FF0000")
                        wb_data.save(path_file)
                        continue

            btn_pilih_rujukan.click()
            time.sleep(1.5)
            
            sheet_data.cell(row=row, column=17).value = tgl_kunjungan
            
            input_tgl = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "txttanggal")))
            input_tgl.clear()
            input_tgl.send_keys(tgl_kunjungan)
            input_tgl.send_keys(Keys.ESCAPE) 
            time.sleep(0.5)
            
            radio_promotif = driver.find_element(By.ID, "tkp50")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", radio_promotif)
            time.sleep(1)
            radio_promotif.click()
            time.sleep(0.5)

            btn_simpan = driver.find_element(By.ID, "btnSimpanPendaftaran")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn_simpan)
            time.sleep(0.5)
            btn_simpan.click()
            tunggu_loading_pace(driver)
            
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//span[@data-notify='message' and contains(text(), 'Data Pendaftaran Berhasil disimpan')]")))
            except: pass
            
            sheet_data.cell(row=row, column=18).value = "SUKSES"
            sheet_data.cell(row=row, column=18).font = Font(color="000000")
            wb_data.save(path_file)
            
            if not yes_to_all:
                print(f"-> Pendaftaran {nama_excel} tersimpan (Tgl: {tgl_kunjungan}).")
                tl = input("Lanjut? (Y=Lanjut / N=Stop / A=Lanjut Semua): ").strip().lower()
                if tl == 'a': yes_to_all = True
                elif tl == 'n': break
            
        except Exception as row_error:
            sheet_data.cell(row=row, column=18).value = f"Gagal sistem: {str(row_error)[:30]}"
            sheet_data.cell(row=row, column=18).font = Font(color="FF0000")
            wb_data.save(path_file)
            
            tl_err = input("Error. Ketik 'n' stop, ENTER/'y' lanjut refresh: ").strip().lower()
            if tl_err == 'n': break
            else:
                driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/EntriDaftarDokkel")
                tunggu_loading_pace(driver)
    else:
        nama_file = os.path.basename(path_file)
        nama_sheet = sheet_data.title
        pesan = format_report_message("SELESAI INPUT PENDAFTARAN PASIEN", nama_file, nama_sheet)
        kirim_notif_group(pesan)


# ====================================================================
# MAIN EKSEKUSI
# ====================================================================

def jalankan_agent():
    path_file, nama_sheet, wb_data = setup_folder_excel()
    sheet_data = wb_data[nama_sheet]
    
    opsi_mode = ["1. Pendaftaran Pasien", "2. Pelayanan Pasien (Input Hasil)"]
    idx_mode, mode_terpilih = menu_interaktif(opsi_mode, "Pilih Mode Operasi Bot:")

    nama_file_excel_user = 'data_user.xlsx' 
    try:
        wb_user = openpyxl.load_workbook(nama_file_excel_user)
        sheet_user = wb_user['Sheet1']
    except Exception as e:
        print(f"Error file user: {e}")
        return

    username = sheet_user['B1'].value
    password = sheet_user['B2'].value

    if not username or not password:
        print("Error: Username/Password kosong di Excel user!")
        return

    print("Membuka browser...")
    chrome_options = Options()
    if ENABLE_KIOSK_PRINTING:
        chrome_options.add_argument("--kiosk-printing")
        if PRINTER_NAME:
            app_state = {
                "recentDestinations": [{"id": PRINTER_NAME, "origin": "local", "account": ""}],
                "selectedDestinationId": PRINTER_NAME,
                "version": 2
            }
            chrome_options.add_experimental_option(
                "prefs",
                {"printing.print_preview_sticky_settings.appState": json.dumps(app_state)}
            )
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    driver.get("https://pcarejkn.bpjs-kesehatan.go.id/eclaim/login")

    print(f"\nMencoba login otomatis untuk: {username}")

    try:
        while True:
            isi_kredensial(driver, username, password)

            print("\n" + "="*60)
            print("MENGUNGGU INPUT CAPTCHA")
            print("1. Username dan Password sudah diisi otomatis.")
            print("2. Silakan ketik kode CAPTCHA di browser (5-6 karakter).")
            print("3. Bot akan klik 'Sign In' otomatis setelah terdeteksi.")
            print("="*60)
            
            if not tunggu_captcha_dan_login(driver):
                print("-> Proses login dibatalkan.")
                return
            
            tunggu_loading_pace(driver)
            hasil_login = tunggu_hasil_login(driver, timeout=15)
            if hasil_login == "success":
                break
            if hasil_login == "failed":
                print("⚠️ Login gagal. Mengulang input username/password...")
                tunggu_loading_pace(driver)
                try:
                    driver.find_element(By.ID, "CaptchaInputText").clear()
                except Exception:
                    pass
                continue
            lanjut = input("Login belum terdeteksi. ENTER = coba lagi, n = batal: ").strip().lower()
            if lanjut == 'n':
                return

        try:
            btn_modal_ok = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.bootbox-accept")))
            btn_modal_ok.click()
            tunggu_loading_pace(driver)
        except: pass
        
        if idx_mode == 0:
            jalankan_pendaftaran(driver, wb_data, sheet_data, path_file)
        elif idx_mode == 1:
            jalankan_pelayanan(driver, wb_data, sheet_data, path_file)

    except Exception as e:
        print(f"-> Terjadi error fatal: {str(e)[:150]}")

    input("\nTekan ENTER untuk menutup browser dan mengakhiri program...")
    driver.quit()

if __name__ == "__main__":
    jalankan_agent()